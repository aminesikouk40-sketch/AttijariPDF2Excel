#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extraire_releve_attijari.py
============================

Extrait les operations (DATE, LIBELLE, DEBIT, CREDIT) d'un releve de compte
bancaire Attijariwafa Bank au format PDF ("RELEVE DE COMPTE BANCAIRE") et
genere un classeur Excel contenant :

  1. Feuille "Releve"        : toutes les operations, dans l'ordre
                                chronologique (DATE | LIBELLE | DEBIT | CREDIT)
                                + ligne de totaux + controle de coherence
                                (solde depart + mouvements = solde final).

  2. Feuille "Credit - Debit" : les operations classees par type, dans deux
                                tableaux places l'un a cote de l'autre :
                                  - a gauche : les operations de CREDIT
                                  - a droite : les operations de DEBIT

Utilisation
-----------
    python3 extraire_releve_attijari.py chemin/vers/releve.pdf [sortie.xlsx]

Si le fichier de sortie n'est pas precise, il est cree a cote du PDF avec
la meme racine de nom et l'extension .xlsx.

Fonctionnement
--------------
Le PDF de ce releve n'est pas un vrai tableau (aucune ligne de separation
verticale) : c'est un texte a colonnes fixes, dans une police a chasse fixe
(Courier-Bold), ou chaque caractere occupe toujours la meme position X sur
la page. Le script repere donc chaque operation par la position horizontale
(en points PDF) de ses caracteres plutot que par les espaces qui les
separent : cela evite les erreurs quand une colonne "deborde" sur la
suivante sans espace (ex. un libelle long colle a la date de valeur).

Ces positions ont ete calibrees sur ce modele de releve Attijariwafa
(colonnes CODE / DATE / LIBELLE / VALEUR / DEBIT / CREDIT). Si un autre
releve utilise une mise en page differente, il faudra recalibrer les
constantes ci-dessous (voir section "CALIBRATION").
"""

import re
import sys
from datetime import date
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CALIBRATION — positions X (en points PDF, origine a gauche de la page)
# des limites de chaque colonne du releve. Reperees empiriquement sur le
# gabarit Attijariwafa "RELEVE DE COMPTE BANCAIRE".
# ---------------------------------------------------------------------------
CODE_X_MAX = 54          # colonne CODE            : x <  54
JOUR_X_MAX = 70          # jour de la DATE          : 54 <= x <  70
MOIS_X_MAX = 89          # mois de la DATE          : 70 <= x <  89
LIBELLE_X_MAX = 300      # LIBELLE                  : 89 <= x < 300
VALEUR_X_MAX = 390       # date de VALEUR           : 300 <= x < 390
MONTANT_SPLIT_X = 505    # DEBIT : 390 <= x < 505 ; CREDIT : x >= 505

ROW_TOLERANCE = 3.0      # tolerance (pts) pour regrouper les caracteres
                          # d'une meme ligne (hauteur de ligne reelle ~12.3pt)

FONT_DONNEES = "Courier-Bold"   # police utilisee pour les lignes d'operations
CODE_RE = re.compile(r"^[0-9A-Z]{6}$")


# ---------------------------------------------------------------------------
# Extraction bas niveau du PDF
# ---------------------------------------------------------------------------
def _dedupe_chars(chars):
    """Le generateur du PDF superpose deux fois la premiere ligne
    d'operation de chaque page de continuation (un caractere espace et le
    vrai caractere empiles exactement aux memes coordonnees). On ne garde
    qu'un caractere par position, en preferant celui qui n'est pas un
    espace."""
    meilleur = {}
    for c in chars:
        cle = (round(c["x0"], 1), round(c["top"], 1))
        if cle not in meilleur or (meilleur[cle]["text"] == " " and c["text"] != " "):
            meilleur[cle] = c
    return list(meilleur.values())


def _cluster_rows(chars, tol=ROW_TOLERANCE):
    """Regroupe les caracteres d'une page en lignes de texte, sur la base
    de leur position verticale ('top'), puis trie chaque ligne de gauche a
    droite."""
    tops = sorted(set(round(c["top"], 1) for c in chars))
    groupes = []
    for t in tops:
        if groupes and abs(t - groupes[-1][-1]) <= tol:
            groupes[-1].append(t)
        else:
            groupes.append([t])
    ligne_de = {}
    for groupe in groupes:
        reference = groupe[0]
        for t in groupe:
            ligne_de[t] = reference
    lignes = {}
    for c in chars:
        cle = ligne_de[round(c["top"], 1)]
        lignes.setdefault(cle, []).append(c)
    return [sorted(lignes[cle], key=lambda c: c["x0"]) for cle in sorted(lignes)]


def _slice_field(row_chars, x_min, x_max):
    """Concatene, dans l'ordre, le texte des caracteres d'une ligne dont la
    position X tombe dans [x_min, x_max)."""
    return "".join(c["text"] for c in row_chars if x_min <= c["x0"] < x_max)


def _parse_montant(texte):
    """Convertit '4 091,83' -> 4091.83. Renvoie None si le champ est vide."""
    texte = texte.strip()
    if not texte:
        return None
    nettoye = texte.replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(nettoye)
    except ValueError:
        return None


def _meilleure_date_operation(jour, mois, date_reference):
    """La colonne DATE du releve ne contient que jour/mois (pas d'annee).
    On choisit, parmi les annees voisines de `date_reference` (la date de
    VALEUR de la meme ligne, ou a defaut une date de reference globale),
    celle qui rapproche le plus la date d'operation de cette reference —
    la date d'operation et la date de valeur d'un releve bancaire ne sont
    jamais tres eloignees l'une de l'autre."""
    candidates = []
    for annee in (date_reference.year - 1, date_reference.year, date_reference.year + 1):
        try:
            candidates.append(date(annee, int(mois), int(jour)))
        except ValueError:
            continue
    if not candidates:
        return None
    return min(candidates, key=lambda d: abs((d - date_reference).days))


def _capter_ligne_resume(texte_page, resume):
    """Repere les lignes 'SOLDE DEPART', 'SOLDE FINAL' et 'TOTAL MOUVEMENTS'
    dans le texte d'une page (hors tableau d'operations) pour pouvoir
    verifier la coherence de l'extraction. Les trois motifs sont
    recherches independamment : ils peuvent apparaitre sur la meme page
    (ex. TOTAL MOUVEMENTS et SOLDE FINAL sont tous deux sur la derniere
    page)."""
    m = re.search(
        r"SOLDE\s+DEPART\s+AU\s+(\d{2})\s+(\d{2})\s+(\d{4}).*?([\d\s]+,\d{2})\s*(CREDITEUR|DEBITEUR)",
        texte_page,
    )
    if m:
        j, mo, a, montant, sens = m.groups()
        resume["solde_depart"] = _parse_montant(montant) * (-1 if sens == "DEBITEUR" else 1)
        resume["date_depart"] = date(int(a), int(mo), int(j))

    m = re.search(
        r"SOLDE\s+FINAL\s+AU\s+(\d{2})\s+(\d{2})\s+(\d{4}).*?([\d\s]+,\d{2})\s*(CREDITEUR|DEBITEUR)",
        texte_page,
    )
    if m:
        j, mo, a, montant, sens = m.groups()
        resume["solde_final"] = _parse_montant(montant) * (-1 if sens == "DEBITEUR" else 1)
        resume["date_final"] = date(int(a), int(mo), int(j))

    m = re.search(r"TOTAL\s+MOUVEMENTS\s+([\d\s]+,\d{2})\s+([\d\s]+,\d{2})", texte_page)
    if m:
        resume["total_debit_pdf"] = _parse_montant(m.group(1))
        resume["total_credit_pdf"] = _parse_montant(m.group(2))


def extraire_operations(pdf_path):
    """Lit le PDF et renvoie (operations, resume).

    operations : liste de dicts {'date': date|None, 'libelle': str,
                                  'debit': float|None, 'credit': float|None}
    resume     : dict avec les cles eventuelles 'solde_depart',
                 'date_depart', 'solde_final', 'date_final',
                 'total_debit_pdf', 'total_credit_pdf' (valeurs lues
                 telles quelles dans le PDF, pour verification).
    """
    brutes = []
    resume = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Les lignes "SOLDE DEPART / SOLDE FINAL / TOTAL MOUVEMENTS" ne
            # sont pas toujours dans la meme police que les lignes
            # d'operations (SOLDE DEPART est en Times-Bold, au milieu de
            # l'en-tete) : on les cherche donc sur le texte complet de la
            # page, toutes polices confondues.
            _capter_ligne_resume(page.extract_text() or "", resume)

            chars = [c for c in page.chars if c["fontname"] == FONT_DONNEES]
            chars = _dedupe_chars(chars)
            if not chars:
                continue
            for row_chars in _cluster_rows(chars):
                code = _slice_field(row_chars, 0, CODE_X_MAX).strip()
                if not CODE_RE.match(code):
                    continue

                jour = _slice_field(row_chars, CODE_X_MAX, JOUR_X_MAX).strip()
                mois = _slice_field(row_chars, JOUR_X_MAX, MOIS_X_MAX).strip()
                if not (jour.isdigit() and mois.isdigit()):
                    continue

                libelle = re.sub(
                    r"\s+", " ", _slice_field(row_chars, MOIS_X_MAX, LIBELLE_X_MAX)
                ).strip()
                valeur_txt = _slice_field(row_chars, LIBELLE_X_MAX, VALEUR_X_MAX).strip()
                debit_txt = _slice_field(row_chars, VALEUR_X_MAX, MONTANT_SPLIT_X).strip()
                credit_txt = _slice_field(row_chars, MONTANT_SPLIT_X, 100000).strip()

                m_valeur = re.match(r"(\d{2})\s+(\d{2})\s+(\d{4})", valeur_txt)
                date_valeur = (
                    date(int(m_valeur.group(3)), int(m_valeur.group(2)), int(m_valeur.group(1)))
                    if m_valeur
                    else None
                )

                brutes.append(
                    {
                        "jour": jour,
                        "mois": mois,
                        "date_valeur": date_valeur,
                        "libelle": libelle,
                        "debit": _parse_montant(debit_txt),
                        "credit": _parse_montant(credit_txt),
                    }
                )

    # Date de reference de secours si une ligne n'a pas de date de valeur
    # exploitable (defensif : n'arrive pas sur ce gabarit).
    reference_globale = resume.get("date_final") or resume.get("date_depart")
    if reference_globale is None:
        premieres_valeurs = [b["date_valeur"] for b in brutes if b["date_valeur"]]
        reference_globale = premieres_valeurs[0] if premieres_valeurs else date.today()

    operations = []
    for b in brutes:
        reference = b["date_valeur"] or reference_globale
        operations.append(
            {
                "date": _meilleure_date_operation(b["jour"], b["mois"], reference),
                "libelle": b["libelle"],
                "debit": b["debit"],
                "credit": b["credit"],
            }
        )

    operations.sort(key=lambda o: (o["date"] is None, o["date"]))
    return operations, resume


# ---------------------------------------------------------------------------
# Construction du classeur Excel
# ---------------------------------------------------------------------------
FONT_NORMAL = Font(name="Arial", size=10)
FONT_GRAS = Font(name="Arial", size=10, bold=True)
FONT_ENTETE = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_TITRE = Font(name="Arial", size=13, bold=True, color="1F3864")

REMPLI_ENTETE = PatternFill("solid", fgColor="1F3864")
REMPLI_CREDIT = PatternFill("solid", fgColor="E2EFDA")
REMPLI_DEBIT = PatternFill("solid", fgColor="FCE4E4")
REMPLI_TOTAL = PatternFill("solid", fgColor="D9D9D9")

BORDURE_FINE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
FORMAT_MONTANT = "#,##0.00"
FORMAT_DATE = "DD/MM/YYYY"


def _style_entete(cell):
    cell.font = FONT_ENTETE
    cell.fill = REMPLI_ENTETE
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDURE_FINE


def _style_cellule(cell, gras=False, montant=False, date_=False, fond=None):
    cell.font = FONT_GRAS if gras else FONT_NORMAL
    cell.border = BORDURE_FINE
    if montant:
        cell.number_format = FORMAT_MONTANT
        cell.alignment = Alignment(horizontal="right")
    elif date_:
        cell.number_format = FORMAT_DATE
        cell.alignment = Alignment(horizontal="center")
    if fond:
        cell.fill = fond


def _feuille_releve(wb, operations, resume):
    ws = wb.active
    ws.title = "Releve"

    ws["A1"] = "RELEVE DES OPERATIONS"
    ws["A1"].font = FONT_TITRE
    ws.merge_cells("A1:D1")

    entetes = ["DATE", "LIBELLE", "DEBIT", "CREDIT"]
    ligne_entete = 3
    for col, texte in enumerate(entetes, start=1):
        c = ws.cell(row=ligne_entete, column=col, value=texte)
        _style_entete(c)

    premiere_ligne = ligne_entete + 1
    for i, op in enumerate(operations):
        r = premiere_ligne + i
        _style_cellule(ws.cell(row=r, column=1, value=op["date"]), date_=True)
        _style_cellule(ws.cell(row=r, column=2, value=op["libelle"]))
        _style_cellule(ws.cell(row=r, column=3, value=op["debit"]), montant=True)
        _style_cellule(ws.cell(row=r, column=4, value=op["credit"]), montant=True)

    derniere_ligne = premiere_ligne + len(operations) - 1
    ligne_total = derniere_ligne + 1
    _style_cellule(ws.cell(row=ligne_total, column=1, value=""), gras=True, fond=REMPLI_TOTAL)
    _style_cellule(ws.cell(row=ligne_total, column=2, value="TOTAL"), gras=True, fond=REMPLI_TOTAL)
    _style_cellule(
        ws.cell(row=ligne_total, column=3, value=f"=SUM(C{premiere_ligne}:C{derniere_ligne})"),
        gras=True, montant=True, fond=REMPLI_TOTAL,
    )
    _style_cellule(
        ws.cell(row=ligne_total, column=4, value=f"=SUM(D{premiere_ligne}:D{derniere_ligne})"),
        gras=True, montant=True, fond=REMPLI_TOTAL,
    )

    # --- Controle de coherence (solde depart + mouvements = solde final) ---
    ligne = ligne_total + 3
    if "solde_depart" in resume:
        ws.cell(row=ligne, column=1, value="Solde depart").font = FONT_GRAS
        c = ws.cell(row=ligne, column=2, value=resume["solde_depart"])
        _style_cellule(c, montant=True)
        ligne += 1
    ligne_solde_depart = ligne - 1

    ws.cell(row=ligne, column=1, value="Total credit").font = FONT_NORMAL
    _style_cellule(ws.cell(row=ligne, column=2, value=f"=D{ligne_total}"), montant=True)
    ligne += 1
    ws.cell(row=ligne, column=1, value="Total debit").font = FONT_NORMAL
    _style_cellule(ws.cell(row=ligne, column=2, value=f"=-C{ligne_total}"), montant=True)
    ligne += 1

    ws.cell(row=ligne, column=1, value="Solde final calcule").font = FONT_GRAS
    if "solde_depart" in resume:
        _style_cellule(
            ws.cell(row=ligne, column=2, value=f"=B{ligne_solde_depart}+D{ligne_total}-C{ligne_total}"),
            gras=True, montant=True,
        )
    ligne += 1

    if "solde_final" in resume:
        ws.cell(row=ligne, column=1, value="Solde final (releve PDF)").font = FONT_NORMAL
        _style_cellule(ws.cell(row=ligne, column=2, value=resume["solde_final"]), montant=True)
        ligne += 1
        ws.cell(row=ligne, column=1, value="Ecart").font = FONT_GRAS
        _style_cellule(
            ws.cell(row=ligne, column=2, value=f"=B{ligne - 2}-B{ligne - 1}"), gras=True, montant=True
        )

    ws.freeze_panes = f"A{premiere_ligne}"
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14


def _bloc_tableau(ws, col_debut, titre, operations, cle_montant, remplissage):
    """Ecrit un tableau DATE / LIBELLE / <montant> a partir de la colonne
    `col_debut` (1 = A, 5 = E, ...) et renvoie le numero de la derniere
    ligne utilisee."""
    c0, c1, c2 = col_debut, col_debut + 1, col_debut + 2

    ws.cell(row=1, column=c0, value=titre).font = FONT_TITRE
    ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c2)

    ligne_entete = 3
    for offset, texte in enumerate(["DATE", "LIBELLE", cle_montant.upper()]):
        c = ws.cell(row=ligne_entete, column=col_debut + offset, value=texte)
        _style_entete(c)

    premiere_ligne = ligne_entete + 1
    for i, op in enumerate(operations):
        r = premiere_ligne + i
        _style_cellule(ws.cell(row=r, column=c0, value=op["date"]), date_=True, fond=remplissage)
        _style_cellule(ws.cell(row=r, column=c1, value=op["libelle"]), fond=remplissage)
        _style_cellule(ws.cell(row=r, column=c2, value=op[cle_montant]), montant=True, fond=remplissage)

    derniere_ligne = premiere_ligne + max(len(operations), 1) - 1
    ligne_total = derniere_ligne + 1
    _style_cellule(ws.cell(row=ligne_total, column=c0, value=""), gras=True, fond=REMPLI_TOTAL)
    _style_cellule(ws.cell(row=ligne_total, column=c1, value="TOTAL"), gras=True, fond=REMPLI_TOTAL)
    if operations:
        formule = f"=SUM({get_column_letter(c2)}{premiere_ligne}:{get_column_letter(c2)}{derniere_ligne})"
    else:
        formule = 0
    _style_cellule(ws.cell(row=ligne_total, column=c2, value=formule), gras=True, montant=True, fond=REMPLI_TOTAL)

    ws.column_dimensions[get_column_letter(c0)].width = 13
    ws.column_dimensions[get_column_letter(c1)].width = 42
    ws.column_dimensions[get_column_letter(c2)].width = 14
    return ligne_total


def _feuille_classement(wb, operations):
    ws = wb.create_sheet("Credit - Debit")

    credits = [o for o in operations if o["credit"] is not None]
    debits = [o for o in operations if o["debit"] is not None]

    _bloc_tableau(ws, 1, "OPERATIONS DE CREDIT", credits, "credit", REMPLI_CREDIT)
    ws.column_dimensions["D"].width = 3  # colonne d'espacement
    _bloc_tableau(ws, 5, "OPERATIONS DE DEBIT", debits, "debit", REMPLI_DEBIT)

    ws.freeze_panes = "A4"


def construire_classeur(operations, resume, chemin_sortie):
    wb = Workbook()
    _feuille_releve(wb, operations, resume)
    _feuille_classement(wb, operations)
    wb.save(chemin_sortie)


# ---------------------------------------------------------------------------
# Programme principal
# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) < 2:
        print("Usage : python3 extraire_releve_attijari.py releve.pdf [sortie.xlsx]")
        sys.exit(1)

    pdf_path = Path(sys.argv[1])
    if not pdf_path.exists():
        print(f"Fichier introuvable : {pdf_path}")
        sys.exit(1)

    sortie = Path(sys.argv[2]) if len(sys.argv) > 2 else pdf_path.with_suffix(".xlsx")

    operations, resume = extraire_operations(str(pdf_path))
    construire_classeur(operations, resume, str(sortie))

    total_debit = sum(o["debit"] for o in operations if o["debit"] is not None)
    total_credit = sum(o["credit"] for o in operations if o["credit"] is not None)

    print(f"{len(operations)} operations extraites -> {sortie}")
    print(f"Total debit  : {total_debit:,.2f}")
    print(f"Total credit : {total_credit:,.2f}")

    if "total_debit_pdf" in resume:
        ok_d = abs(total_debit - resume["total_debit_pdf"]) < 0.01
        ok_c = abs(total_credit - resume["total_credit_pdf"]) < 0.01
        print(
            f"Controle vs totaux imprimes sur le releve : "
            f"debit {'OK' if ok_d else 'ECART !'} / credit {'OK' if ok_c else 'ECART !'}"
        )


if __name__ == "__main__":
    main()
